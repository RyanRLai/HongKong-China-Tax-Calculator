from __future__ import annotations

from decimal import Decimal
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - exercised only before dependencies are installed.
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

from models import EvidenceRecord, IncomeRecord, ParsedDocumentLog, TaxResult
from reporting.calculation_process import build_calculation_process_text
from tax.rules import TaxRules


HEADER_FILL = PatternFill("solid", fgColor="1F4E78") if PatternFill else None
HEADER_FONT = Font(color="FFFFFF", bold=True) if Font else None
MONEY_FMT = '#,##0.00'
RATE_FMT = '0.000000'
DATE_FMT = 'yyyy-mm-dd'


def write_excel_report(
    output_path: Path,
    records: list[IncomeRecord],
    tax_results: list[TaxResult],
    logs: list[ParsedDocumentLog],
    evidence: dict[str, EvidenceRecord],
    rules: TaxRules,
) -> None:
    if Workbook is None:
        raise RuntimeError("Missing dependency: install openpyxl with `python3 -m pip install -r requirements.txt`.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    result_by_id = {result.record_id: result for result in tax_results}
    _write_summary(workbook, records, tax_results)
    _write_chinese_summary(workbook, records, tax_results, rules)
    _write_income_detail(workbook, records, result_by_id)
    _write_calculation_process(workbook, records, result_by_id, rules)
    _write_dividend_credit(workbook, records, result_by_id, evidence)
    _write_parsing_log(workbook, logs)
    _write_hkex_evidence(workbook, evidence)
    _write_review_flags(workbook, records, tax_results, evidence)
    _write_fx_rates(workbook, rules)
    _write_sources_assumptions(workbook, rules)

    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    workbook.save(output_path)


def _write_summary(workbook: Workbook, records: list[IncomeRecord], results: list[TaxResult]) -> None:
    ws = workbook.create_sheet("Summary")
    ws.append(["Metric", "Value", "Notes"])
    total_income = sum((r.taxable_amount_rmb for r in results), Decimal("0"))
    total_tax = sum((r.mainland_tax_rmb for r in results), Decimal("0"))
    total_credit = sum((r.allowable_credit_rmb for r in results), Decimal("0"))
    total_due = sum((r.estimated_tax_due_rmb for r in results), Decimal("0"))
    review_count = sum(1 for result in results if result.review_flags)
    ws.append(["Records parsed", len(records), "Income rows parsed from bank statement PDFs"])
    ws.append(["Taxable income RMB", total_income, "Uses FX_Rates placeholders unless updated"])
    ws.append(["Mainland IIT before credit RMB", total_tax, "Default rate comes from tax_rules"])
    ws.append(["Allowable foreign tax credit RMB", total_credit, "Only confirmed foreign income tax should be used"])
    ws.append(["Estimated tax due RMB", total_due, "Before professional/manual review"])
    ws.append(["Rows needing review", review_count, "See Review_Flags"])
    for row in range(3, 7):
        ws.cell(row=row, column=2).number_format = MONEY_FMT


def _write_chinese_summary(
    workbook: Workbook,
    records: list[IncomeRecord],
    results: list[TaxResult],
    rules: TaxRules,
) -> None:
    ws = workbook.create_sheet("Chinese_Summary")
    ws.append(["项目", "数值", "说明"])
    total_income = sum((r.taxable_amount_rmb for r in results), Decimal("0"))
    total_tax = sum((r.mainland_tax_rmb for r in results), Decimal("0"))
    total_credit = sum((r.allowable_credit_rmb for r in results), Decimal("0"))
    total_due = sum((r.estimated_tax_due_rmb for r in results), Decimal("0"))
    review_count = sum(1 for result in results if result.review_flags)
    ws.append(["税务年份", rules.year, "读取 tax_rules 中对应年份的规则"])
    ws.append(["解析出的收入记录数", len(records), "从银行 PDF 中识别出的收入行"])
    ws.append(["人民币应税收入合计", total_income, "每条收入按 FX_Rates 汇率换算后相加"])
    ws.append(["中国内地税额合计", total_tax, "人民币应税收入 × 税率"])
    ws.append(["可抵扣境外税额合计", total_credit, "只使用已经确认的境外扣税金额"])
    ws.append(["估算应缴税额合计", total_due, "中国内地税额 - 可抵扣境外税额"])
    ws.append(["需要人工复核的行数", review_count, "请查看 Review_Flags 工作表"])
    ws.append(["重要提醒", "仅供学习和整理使用", "正式申报前请向专业人士或税务机关确认"])
    for row in range(4, 8):
        ws.cell(row=row, column=2).number_format = MONEY_FMT


def _write_income_detail(workbook: Workbook, records: list[IncomeRecord], result_by_id: dict[str, TaxResult]) -> None:
    ws = workbook.create_sheet("Income_Detail")
    ws.append([
        "record_id",
        "date",
        "tax_year",
        "bank",
        "account_last4",
        "income_type",
        "stock_code",
        "company_name",
        "currency",
        "net_amount_original",
        "taxable_amount_original",
        "fx_rate",
        "taxable_amount_rmb",
        "mainland_tax_rmb",
        "source_file",
        "page",
        "confidence",
        "review_flags",
    ])
    for record in records:
        result = result_by_id[record.record_id]
        ws.append([
            record.record_id,
            record.transaction_date,
            record.tax_year,
            record.bank,
            record.account_last4,
            record.income_type,
            record.stock_code,
            record.company_name,
            record.currency,
            record.net_amount,
            result.taxable_amount_original,
            result.fx_rate,
            result.taxable_amount_rmb,
            result.mainland_tax_rmb,
            record.source_file,
            record.source_page,
            record.confidence,
            "; ".join(result.review_flags),
        ])


def _write_calculation_process(
    workbook: Workbook,
    records: list[IncomeRecord],
    result_by_id: dict[str, TaxResult],
    rules: TaxRules,
) -> None:
    ws = workbook.create_sheet("Calculation_Process")
    ws.append([
        "record_id",
        "date",
        "income_type",
        "currency",
        "taxable_amount_original",
        "fx_rate",
        "taxable_amount_rmb",
        "mainland_tax_rmb",
        "candidate_foreign_tax_original",
        "candidate_foreign_tax_rmb",
        "allowable_credit_rmb",
        "estimated_tax_due_rmb",
        "calculation_process_cn",
        "review_flags",
    ])
    for record in records:
        result = result_by_id[record.record_id]
        ws.append([
            record.record_id,
            record.transaction_date,
            record.income_type,
            record.currency,
            result.taxable_amount_original,
            result.fx_rate,
            result.taxable_amount_rmb,
            result.mainland_tax_rmb,
            result.candidate_foreign_tax_original,
            result.candidate_foreign_tax_rmb,
            result.allowable_credit_rmb,
            result.estimated_tax_due_rmb,
            build_calculation_process_text(record, result, rules),
            "; ".join(result.review_flags),
        ])


def _write_dividend_credit(
    workbook: Workbook,
    records: list[IncomeRecord],
    result_by_id: dict[str, TaxResult],
    evidence: dict[str, EvidenceRecord],
) -> None:
    ws = workbook.create_sheet("Dividend_Credit")
    ws.append([
        "record_id",
        "date",
        "stock_code",
        "company_name",
        "currency",
        "net_received",
        "candidate_foreign_tax",
        "credit_limit_rmb",
        "allowable_credit_rmb",
        "estimated_tax_due_rmb",
        "hkex_status",
        "hkex_withholding_rate",
        "review_flags",
    ])
    for record in records:
        if record.income_type != "dividend":
            continue
        result = result_by_id[record.record_id]
        ev = evidence.get(record.record_id)
        ws.append([
            record.record_id,
            record.transaction_date,
            record.stock_code,
            record.company_name,
            record.currency,
            record.net_amount,
            result.candidate_foreign_tax_original,
            result.credit_limit_rmb,
            result.allowable_credit_rmb,
            result.estimated_tax_due_rmb,
            ev.status if ev else "not_searched",
            ev.withholding_rate if ev and ev.withholding_rate is not None else "",
            "; ".join(result.review_flags + ((("; ".join(ev.review_flags),) if ev else tuple()))),
        ])


def _write_parsing_log(workbook: Workbook, logs: list[ParsedDocumentLog]) -> None:
    ws = workbook.create_sheet("Bank_Parsing_Log")
    ws.append(["source_file", "bank", "status", "pages", "text_chars", "records_found", "warnings"])
    for log in logs:
        ws.append([log.source_file, log.bank, log.status, log.pages, log.text_chars, log.records_found, "; ".join(log.warnings)])


def _write_hkex_evidence(workbook: Workbook, evidence: dict[str, EvidenceRecord]) -> None:
    ws = workbook.create_sheet("HKEX_Evidence")
    ws.append([
        "record_id",
        "source",
        "status",
        "url",
        "announcement_title",
        "announcement_date",
        "withholding_rate",
        "matched_text",
        "review_flags",
    ])
    for ev in evidence.values():
        ws.append([
            ev.record_id,
            ev.source,
            ev.status,
            ev.url,
            ev.announcement_title,
            ev.announcement_date,
            ev.withholding_rate if ev.withholding_rate is not None else "",
            ev.matched_text,
            "; ".join(ev.review_flags),
        ])


def _write_review_flags(
    workbook: Workbook,
    records: list[IncomeRecord],
    results: list[TaxResult],
    evidence: dict[str, EvidenceRecord],
) -> None:
    ws = workbook.create_sheet("Review_Flags")
    ws.append(["record_id", "date", "income_type", "stock_code", "source_file", "page", "flag", "suggested_action"])
    record_by_id = {record.record_id: record for record in records}
    actions = {
        "withholding_not_confirmed": "Check bank statement and HKEX announcement before claiming credit",
        "no_confirmed_foreign_tax_credit": "Do not claim tax credit until withholding evidence is available",
        "fx_rate_missing_or_placeholder": "Fill official RMB central parity rate in FX_Rates",
        "manual_hkex_announcement_match_required": "Search HKEXnews by stock code and dividend date",
        "currency_not_found": "Confirm the currency from the bank statement section",
    }
    for result in results:
        record = record_by_id[result.record_id]
        flags = list(result.review_flags)
        ev = evidence.get(result.record_id)
        if ev:
            flags.extend(ev.review_flags)
        for flag in dict.fromkeys(flags):
            ws.append([
                record.record_id,
                record.transaction_date,
                record.income_type,
                record.stock_code,
                record.source_file,
                record.source_page,
                flag,
                actions.get(flag, "Manual review required"),
            ])


def _write_fx_rates(workbook: Workbook, rules: TaxRules) -> None:
    ws = workbook.create_sheet("FX_Rates")
    ws.append(["currency", "fx_rate_to_rmb", "review_status", "notes"])
    for currency in sorted(rules.fx_rates):
        status = "placeholder_review_required" if rules.fx_rates[currency] == Decimal("1") else "configured_review_required"
        ws.append([currency, rules.fx_rates[currency], status, "Update with official RMB central parity rate before filing"])


def _write_sources_assumptions(workbook: Workbook, rules: TaxRules) -> None:
    ws = workbook.create_sheet("Sources_Assumptions")
    ws.append(["Topic", "Detail"])
    ws.append(["Tax year", rules.year])
    ws.append(["Mainland IIT rate", rules.mainland_rate])
    ws.append(["Scope", "China resident individual, Hong Kong bank statements, interest/dividend income"])
    ws.append(["Official China tax source", "Individual Income Tax Law and overseas income credit rules from China Tax authorities"])
    ws.append(["HKEX source", "HKEXnews issuer announcements should be used as dividend evidence"])
    ws.append(["Safety", "PDFs remain local; report masks account data to last four digits"])
    ws.append(["Limitation", "This workbook is for organization and estimation, not a substitute for professional tax advice"])
    for note in rules.notes:
        ws.append(["Rule note", note])


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if hasattr(cell.value, "as_tuple"):
                cell.number_format = MONEY_FMT
            if cell.column_letter.lower().endswith("b") and "date" in str(ws.cell(1, cell.column).value).lower():
                cell.number_format = DATE_FMT
            header = str(ws.cell(1, cell.column).value).lower()
            if "rate" in header:
                cell.number_format = RATE_FMT
            if "amount" in header or "tax" in header or "credit" in header or "income" in header:
                if hasattr(cell.value, "as_tuple"):
                    cell.number_format = MONEY_FMT
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 48)
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = max(width, 10)
