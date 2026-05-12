from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import Mock, patch

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from models import IncomeRecord, TaxResult
from parsers.base import ParseResult
from reporting.calculation_process import build_calculation_process_text
from tax import TaxRules


def make_record() -> IncomeRecord:
    return IncomeRecord(
        record_id="r1",
        tax_year=2025,
        bank="fake_bank",
        account_last4="1234",
        transaction_date=date(2025, 8, 1),
        income_type="dividend",
        currency="HKD",
        net_amount=Decimal("80.00"),
        gross_amount=Decimal("100.00"),
        withholding_tax=Decimal("10.00"),
        source_file="fake.pdf",
        source_page=1,
        source_line="Cash Dividend",
        confidence=0.9,
    )


def make_result(review_flags: tuple[str, ...] = ()) -> TaxResult:
    return TaxResult(
        record_id="r1",
        taxable_amount_original=Decimal("100.00"),
        fx_rate=Decimal("1.10"),
        taxable_amount_rmb=Decimal("110.00"),
        mainland_tax_rmb=Decimal("22.00"),
        candidate_foreign_tax_original=Decimal("10.00"),
        candidate_foreign_tax_rmb=Decimal("11.00"),
        credit_limit_rmb=Decimal("22.00"),
        allowable_credit_rmb=Decimal("11.00"),
        estimated_tax_due_rmb=Decimal("11.00"),
        review_flags=review_flags,
    )


class ReportingCalculationProcessTests(unittest.TestCase):
    def test_calculation_process_text_explains_each_formula(self):
        text = build_calculation_process_text(
            make_record(),
            make_result(),
            TaxRules(year=2025, mainland_rate=Decimal("0.20"), fx_rates={"HKD": Decimal("1.10")}, notes=()),
        )

        self.assertIn("应税收入人民币 = 原币应税金额 × 汇率 = 100.00 × 1.10 = 110.00 RMB", text)
        self.assertIn("中国内地税额 = 应税收入人民币 × 税率 = 110.00 × 0.20 = 22.00 RMB", text)
        self.assertIn("估算应缴税 = 中国内地税额 - 可抵扣境外税额 = 22.00 - 11.00 = 11.00 RMB", text)


class WorkflowTests(unittest.TestCase):
    def test_run_tax_calculation_returns_summary(self):
        import workflow

        parser = Mock()
        parser.parse_directory.return_value = ParseResult(records=[make_record()], logs=[Mock()])
        result = make_result(review_flags=("needs_review",))
        evidence_provider = Mock()
        evidence_provider.lookup_many.return_value = {}

        with (
            patch.object(workflow, "get_parser", return_value=parser),
            patch.object(
                workflow,
                "load_tax_rules",
                return_value=TaxRules(year=2025, mainland_rate=Decimal("0.20"), fx_rates={"HKD": Decimal("1")}, notes=()),
            ),
            patch.object(workflow, "HKEXEvidenceProvider", return_value=evidence_provider),
            patch.object(workflow, "calculate_tax_results", return_value=[result]),
            patch.object(workflow, "write_excel_report") as write_report,
        ):
            summary = workflow.run_tax_calculation(
                year=2025,
                bank="standard_chartered_hk",
                input_dir=Path("input"),
                output_path=Path("output.xlsx"),
                rules_dir=Path("tax_rules"),
                hkex_enabled=False,
            )

        self.assertEqual(summary.parsed_pdfs, 1)
        self.assertEqual(summary.income_records, 1)
        self.assertEqual(summary.rows_needing_review, 1)
        self.assertEqual(summary.output_path, Path("output.xlsx"))
        write_report.assert_called_once()


@unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
class ExcelWorkbookTests(unittest.TestCase):
    def test_excel_report_contains_chinese_and_calculation_sheets(self):
        from openpyxl import load_workbook
        from reporting import write_excel_report

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "report.xlsx"
            write_excel_report(
                output_path=output_path,
                records=[make_record()],
                tax_results=[make_result()],
                logs=[],
                evidence={},
                rules=TaxRules(year=2025, mainland_rate=Decimal("0.20"), fx_rates={"HKD": Decimal("1.10")}, notes=()),
            )

            workbook = load_workbook(output_path)

        self.assertIn("Chinese_Summary", workbook.sheetnames)
        self.assertIn("Calculation_Process", workbook.sheetnames)
        process_sheet = workbook["Calculation_Process"]
        self.assertIn("估算应缴税", process_sheet["M2"].value)


if __name__ == "__main__":
    unittest.main()
